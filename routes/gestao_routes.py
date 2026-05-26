import logging
from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, send_file, Response
from flask_login import login_required, current_user
from models import db, Retorno, Sessao
from datetime import datetime, date
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from utils import parse_date, classificar_prazo, prazo_envio, dias_atraso, PRAZO_CORES, PRAZO_ORDEM, registrar_auditoria

logger = logging.getLogger(__name__)
gestao_bp = Blueprint("core", __name__)


def _get_retornos_data(incluir_excluidos=False):
    q = Retorno.query
    if not incluir_excluidos:
        q = q.filter(Retorno.status_exclusao != "Excluído")
    retornos = q.order_by(Retorno.id.desc()).all()
    return [r.to_dict() for r in retornos]


# ─── Dashboard ───────────────────────────────────────────────────────────────

@gestao_bp.route("/")
@gestao_bp.route("/dashboard")
@login_required
def dashboard():
    retornos = _get_retornos_data()
    for r in retornos:
        r["_prazo"] = classificar_prazo(r)
        r["_dias"] = dias_atraso(r)

    from config import Config
    STATUS_FLUXO = Config.STATUS_FLUXO

    # KPIs
    n_andamento = sum(1 for r in retornos if r["status_retorno"] != STATUS_FLUXO[4])
    n_prioritarios = sum(1 for r in retornos if r["status_retorno"] != STATUS_FLUXO[4]
                         and r.get("tipo_retorno", "").lower() == "prioritário")
    kpis = {
        "andamento": n_andamento,
        "prioritarios": n_prioritarios,
        "aguardando_avaliacao": sum(1 for r in retornos if r["status_retorno"] == STATUS_FLUXO[0]),
        "aguardando_execucao": sum(1 for r in retornos if r["status_retorno"] == STATUS_FLUXO[1]),
        "aguardando_validacao": sum(1 for r in retornos if r["status_retorno"] == STATUS_FLUXO[2]),
        "aguardando_envio": sum(1 for r in retornos if r["status_retorno"] == STATUS_FLUXO[3]),
        "concluidos": sum(1 for r in retornos if r["status_retorno"] == STATUS_FLUXO[4]),
        "atrasados": sum(1 for r in retornos if r.get("_prazo") == "Em Atraso"),
    }

    # Pizza data
    pizza_data = {}
    for p in PRAZO_ORDEM:
        n = sum(1 for r in retornos if r["_prazo"] == p and r["_prazo"] != "Concluído")
        if n > 0:
            pizza_data[p] = n

    # Ranking: prioridades customizadas (prazo_execucao_customizado preenchido) no TOPO
    ranking = []
    retornos_abertos = [r for r in retornos if r["status_retorno"] != "Concluído" and r["_prazo"] != "Concluído"]

    def _is_custom_priority(row):
        val = str(row.get("prazo_execucao_customizado", "") or "").strip()
        return bool(val and val not in ("", "-", "nan"))

    # Separa em dois grupos
    prio = [r for r in retornos_abertos if _is_custom_priority(r)]
    normal = [r for r in retornos_abertos if not _is_custom_priority(r)]
    # Ordena: prioridade por prazo_custom asc, normal por dias desc
    prio.sort(key=lambda r: parse_date(str(r.get("prazo_execucao_customizado", ""))) or date.max)
    normal.sort(key=lambda r: r["_dias"], reverse=True)
    # Junta: prioridades primeiro
    sorted_retornos = prio + normal

    for i, r in enumerate(sorted_retornos[:20]):
        ranking.append({
            "pos": i + 1,
            "proposta": r.get("numero_proposta", ""),
            "projeto": r.get("nome_projeto", ""),
            "cliente": r.get("nome_cliente", ""),
            "dias": r["_dias"],
            "prazo": r["_prazo"],
            "is_prio": _is_custom_priority(r),
            "prazo_custom": r.get("prazo_execucao_customizado", ""),
            "ticket": r.get("ticket_atendimento", ""),
        })

    # Setor chart
    from config import Config
    setor_data = {}
    for r in retornos:
        if r["status_retorno"] == "Concluído" or r["_prazo"] == "Concluído":
            continue
        area = r.get("area_responsavel", "").strip()
        if area:
            setor_data[area] = setor_data.get(area, 0) + 1

    # Dashboard table: top 20 critical retornos (not concluded, ordered by urgency)
    from sqlalchemy import case
    # Priority: atrasados first (prazo = Em Atraso), then prioridade, then most days open
    retornos_q = Retorno.query.filter(
        Retorno.status_exclusao != "Excluído",
        Retorno.status_retorno != "Concluído"
    ).order_by(
        Retorno.status_retorno.asc(),
        Retorno.tipo_retorno.desc(),
        Retorno.id.desc()
    ).limit(20).all()
    retornos_dashboard = [r.to_dict() for r in retornos_q]
    for r in retornos_dashboard:
        r["_prazo"] = classificar_prazo(r)
        r["_dias"] = dias_atraso(r)
    # Re-sort by urgency: atrasados first, then by days desc
    retornos_dashboard.sort(key=lambda r: (
        0 if r["_prazo"] == "Em Atraso" else 1,
        -r["_dias"]
    ))

    return render_template("dashboard.html", kpis=kpis, pizza_data=pizza_data,
                           ranking=ranking, setor_data=setor_data,
                           PRAZO_CORES=PRAZO_CORES, PRAZO_ORDEM=PRAZO_ORDEM,
                           STATUS_FLUXO=STATUS_FLUXO,
                           retornos_dashboard=retornos_dashboard)


# ─── Lista de Retornos ───────────────────────────────────────────────────────

@gestao_bp.route("/lista")
@login_required
def lista():
    page = request.args.get("page", 1, type=int)
    per_page = 50
    incluir = request.args.get("excluidos", "0") == "1"

    from config import Config

    status_filter = request.args.get("status", "")
    tipo_filter = request.args.get("tipo", "")
    area_filter = request.args.get("area", "")
    prazo_filter = request.args.get("prazo", "")
    search = request.args.get("search", "").strip().lower()

    q = Retorno.query
    if not incluir:
        q = q.filter(Retorno.status_exclusao != "Excluído")

    if status_filter:
        q = q.filter(Retorno.status_retorno == status_filter)
    if tipo_filter:
        q = q.filter(Retorno.tipo_retorno == tipo_filter)
    if area_filter:
        q = q.filter(Retorno.area_responsavel == area_filter)

    q = q.order_by(Retorno.id.desc())

    from math import ceil
    from sqlalchemy import or_

    if prazo_filter == "urgente":
        all_ret = [r.to_dict() for r in q.all()]
        for r in all_ret:
            r["_prazo"] = classificar_prazo(r)
        filtered = [r for r in all_ret if r["_prazo"] in ("Em Atraso", "Prazo Final")]
        total = len(filtered)
        total_pages = max(1, ceil(total / per_page))
        start = (page - 1) * per_page
        retornos = filtered[start:start + per_page]
    else:
        if search:
            search_filter = or_(
                Retorno.ticket_atendimento.ilike(f"%{search}%"),
                Retorno.nome_projeto.ilike(f"%{search}%"),
                Retorno.nome_cliente.ilike(f"%{search}%"),
                Retorno.nome_pmo.ilike(f"%{search}%"),
                Retorno.codigo_documento.ilike(f"%{search}%"),
                Retorno.area_responsavel.ilike(f"%{search}%"),
                Retorno.status_retorno.ilike(f"%{search}%"),
            )
            q = q.filter(search_filter)

        pagination = q.paginate(page=page, per_page=per_page, error_out=False)
        retornos = [r.to_dict() for r in pagination.items]
        total = pagination.total
        total_pages = pagination.pages

    # Add prazo classification for display
    for r in retornos:
        r["_prazo"] = classificar_prazo(r)

    # Get distinct areas for filter dropdown
    areas_q = Retorno.query.with_entities(Retorno.area_responsavel).distinct()
    if not incluir:
        areas_q = areas_q.filter(Retorno.status_exclusao != "Excluído")
    areas = sorted(set(r[0] for r in areas_q.all() if r[0]))

    return render_template("lista_retornos.html",
        retornos=retornos, areas=areas, incluir=incluir,
        status_filter=status_filter, tipo_filter=tipo_filter,
        area_filter=area_filter, search=request.args.get("search", ""),
        pode_excluir=current_user.pode_excluir,
        page=page, total_pages=total_pages, total=total,
        prazo_filter=prazo_filter,
    )


@gestao_bp.route("/detalhes/<path:ticket>")
@login_required
def detalhes(ticket):
    ret = Retorno.query.filter_by(ticket_atendimento=ticket).first()
    if not ret:
        flash("Retorno não encontrado.", "danger")
        return redirect(url_for("core.lista"))
    return render_template("detalhes.html", r=ret.to_dict())


@gestao_bp.route("/detalhes/<path:ticket>/pdf")
@login_required
def relatorio_pdf(ticket):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from xml.sax.saxutils import escape

    ret = Retorno.query.filter_by(ticket_atendimento=ticket).first()
    if not ret:
        flash("Retorno não encontrado.", "danger")
        return redirect(url_for("core.lista"))
    r = ret.to_dict()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    azul = HexColor("#1a2738")
    azul_claro = HexColor("#4472C4")
    cinza = HexColor("#6b7280")

    def e(v):
        return escape(str(v or ""))

    def est(nome, **kw):
        defaults = dict(fontName="Helvetica", fontSize=9, textColor=HexColor("#000000"), spaceAfter=2)
        defaults.update(kw)
        return ParagraphStyle(nome, **defaults)

    elementos = []

    # Cabeçalho
    elementos.append(Paragraph("SARA - Relatorio de Retorno", est("Titulo", fontName="Helvetica-Bold", fontSize=20, textColor=azul, spaceBefore=0, spaceAfter=16)))
    cab = f"Ticket: {e(r.get('ticket_atendimento'))}     ID: {e(r.get('id_retorno'))}     {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    elementos.append(Paragraph(cab, est("Sub", fontSize=9, textColor=cinza, spaceAfter=12)))
    elementos.append(HRFlowable(width="100%", color=azul_claro, thickness=2, spaceAfter=14))

    # Seções
    secoes = [
        ("ETAPA 1 - SOLICITAÇÃO", [
            "ID", "ticket_atendimento", "Data Receb. GP", "gp_recebimento_cliente",
            "Data Envio GP/CE", "gp_ce_envio_recebimento", "PMO", "nome_pmo",
            "N Proposta", "numero_proposta", "Cliente", "nome_cliente",
            "Projeto", "nome_projeto", "Cod. Documento", "codigo_documento",
            "Tipo", "tipo_retorno", "Justificativa", "justificativa_prioridade",
            "Status", "status_retorno",
        ]),
        ("ETAPA 2 - AVALIACAO CE", [
            "Avaliador", "avaliador_retorno", "Area Resp.", "area_responsavel",
            "Data Envio CE/AR", "ce_ar_envio_recebimento", "Obs.", "obs_avaliacao_ce",
        ]),
        ("ETAPA 3 - EXECUCAO", [
            "Responsavel", "responsavel_execucao", "Status Exec.", "status_execucao",
            "Data Envio AR/CE", "ar_ce_envio_recebimento", "Comentario", "comentario_execucao",
        ]),
        ("ETAPA 4 - VALIDACAO CE", [
            "Aprovador", "aprovador_retorno", "Status CE", "status_ce",
            "Data Envio CE/GP", "ce_gp_envio_recebimento", "Comentario", "comentario_validacao",
        ]),
        ("ETAPA 5 - ENVIO", [
            "Data Envio GP/Cliente", "gp_envio_cliente", "Obs.", "obs_envio",
        ]),
    ]

    for titulo, campos in secoes:
        elementos.append(Paragraph(titulo, est("Sec", fontName="Helvetica-Bold", fontSize=11, textColor=HexColor("#ffffff"), spaceBefore=16, spaceAfter=6)))
        dados = []
        for i in range(0, len(campos), 2):
            label = campos[i]
            chave = campos[i+1]
            val = str(r.get(chave, "") or "").strip()
            if val in ("-", "", "nan"):
                val = "-"
            dados.append([f"{label}:", val])

        t = Table(dados, colWidths=[doc.width * 0.3, doc.width * 0.7])
        t.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#e0e0e0")),
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#f5f7fa")),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [HexColor("#fafafa"), HexColor("#ffffff")]),
        ]))
        elementos.append(t)
        elementos.append(Spacer(1, 6))

    # Exclusao
    if r.get("status_exclusao") == "Excluído":
        elementos.append(HRFlowable(width="100%", color=HexColor("#C00000"), thickness=1, spaceAfter=4))
        elementos.append(Paragraph("EXCLUSAO", est("Exc", fontName="Helvetica-Bold", fontSize=10, textColor=HexColor("#C00000"), spaceBefore=4, spaceAfter=2)))
        exc = [
            ["Status:", e(r.get("status_exclusao"))],
            ["Responsavel:", e(r.get("responsavel_exclusao"))],
            ["Data:", e(r.get("data_exclusao"))],
            ["Motivo:", e(r.get("detalhamento_exclusao"))],
        ]
        t = Table(exc, colWidths=[doc.width * 0.3, doc.width * 0.7])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#f5c6cb")),
            ("BACKGROUND", (0, 0), (-1, -1), HexColor("#fff5f5")),
        ]))
        elementos.append(t)

    elementos.append(Spacer(1, 16))
    rodape = f"Documento gerado pelo SARA em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Usuario: {e(current_user.nome_exibicao)}"
    elementos.append(HRFlowable(width="100%", color=HexColor("#d0d0d0"), thickness=0.5, spaceAfter=4))
    elementos.append(Paragraph(rodape, est("Rod", fontSize=7, textColor=cinza, alignment=1)))

    try:
        doc.build(elementos)
    except Exception as exc:
        flash(f"Erro ao gerar PDF: {exc}", "danger")
        return redirect(url_for("core.detalhes", ticket=ticket))

    buf.seek(0)
    nome_arquivo = f"retorno_{r.get('ticket_atendimento', 'ticket')}.pdf".replace("/", "_")
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True,
                     download_name=nome_arquivo)


@gestao_bp.route("/excluir/<path:ticket>", methods=["POST"])
@login_required
def excluir(ticket):
    if not current_user.pode_excluir:
        flash("Sem permissão.", "danger")
        return redirect(url_for("core.lista"))

    ret = Retorno.query.filter_by(ticket_atendimento=ticket).first()
    if not ret:
        flash("Retorno não encontrado.", "danger")
        return redirect(url_for("core.lista"))

    motivo = request.form.get("motivo", "").strip()
    if not motivo:
        flash("Informe o motivo da exclusão.", "warning")
        return redirect(url_for("core.detalhes", ticket=ticket))

    ret.status_exclusao = "Excluído"
    ret.responsavel_exclusao = current_user.nome_exibicao
    ret.detalhamento_exclusao = motivo
    ret.data_exclusao = datetime.now().strftime("%d/%m/%Y")
    registrar_auditoria(acao="excluir_retorno", entidade="retorno", ticket_retorno=ticket,
                        status_anterior="Ativo", status_novo="Excluído",
                        detalhes=f"Responsavel: {current_user.nome_exibicao}, Motivo: {motivo[:200]}")
    db.session.commit()
    logger.info("Exclusao logica: ticket=%s user=%s motivo=%s", ticket, current_user.usuario, motivo[:50])
    flash(f"Retorno {ticket} marcado como Excluído.", "success")
    return redirect(url_for("core.lista"))


@gestao_bp.route("/exportar", methods=["GET"])
@login_required
def exportar():
    retornos = _get_retornos_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Retornos"

    from config import Config as Cfg
    labels = Cfg.COL_LABELS

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
    hdr_aln = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D0D8E8")
    hdr_brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_keys = list(labels.keys())
    for ci, col_name in enumerate(labels.values(), 1):
        c = ws.cell(row=1, column=ci, value=col_name)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = hdr_aln; c.border = hdr_brd

    alt_fill = PatternFill("solid", fgColor="EEF2F9")
    norm_fill = PatternFill("solid", fgColor="FFFFFF")
    d_font = Font(size=9, name="Segoe UI")
    d_brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ri, r in enumerate(retornos):
        bg = alt_fill if ri % 2 == 0 else norm_fill
        for ci, col in enumerate(col_keys, 1):
            val = str(r.get(col, ""))
            c = ws.cell(row=ri + 2, column=ci, value=val if val not in ("-", "nan") else "")
            c.fill = bg; c.font = d_font; c.border = d_brd

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="retornos.xlsx")


# ─── Edição Manual ───────────────────────────────────────────────────────────

@gestao_bp.route("/edicao", methods=["GET", "POST"])
@login_required
def edicao():
    if not current_user.pode_editar:
        flash("Sem permissão para edição manual.", "danger")
        return redirect(url_for("core.dashboard"))

    if request.method == "POST":
        ticket = request.form.get("ticket", "").strip()
        if not ticket:
            flash("Selecione um ticket.", "warning")
            return redirect(url_for("core.edicao"))

        ret = Retorno.query.filter_by(ticket_atendimento=ticket).first()
        if not ret:
            flash("Ticket não encontrado.", "danger")
            return redirect(url_for("core.edicao"))

        campos_permitidos = {
            "gp_recebimento_cliente", "nome_pmo", "numero_proposta", "nome_cliente",
            "nome_projeto", "codigo_documento", "tipo_retorno", "justificativa_prioridade",
            "gp_ce_envio_recebimento", "avaliador_retorno", "area_responsavel",
            "ce_ar_envio_recebimento", "obs_avaliacao_ce", "status_execucao",
            "responsavel_execucao", "ar_ce_envio_recebimento", "comentario_execucao",
            "aprovador_retorno", "status_ce", "ce_gp_envio_recebimento", "comentario_validacao",
            "gp_envio_cliente", "obs_envio", "status_retorno", "prazo_execucao_customizado",
        }
        try:
            for key in request.form:
                if key in campos_permitidos:
                    setattr(ret, key, request.form[key].strip())
            db.session.commit()
            flash(f"Ticket {ticket} atualizado.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao salvar: {e}", "danger")
        return redirect(url_for("core.edicao", ticket=ticket))

    ticket_sel = request.args.get("ticket", "")
    record = None
    if ticket_sel:
        ret = Retorno.query.filter_by(ticket_atendimento=ticket_sel).first()
        if ret:
            record = ret.to_dict()
            # Preenche data de recebimento com data de criação se vazio
            if not record.get("gp_recebimento_cliente") or record["gp_recebimento_cliente"].strip() in ("", "-"):
                record["gp_recebimento_cliente"] = ret.criado_em.strftime("%d/%m/%Y") if ret.criado_em else ""

    retornos = Retorno.query.filter(Retorno.status_exclusao != "Excluído").order_by(Retorno.id.desc()).all()
    tickets = [r.ticket_atendimento for r in retornos]
    return render_template("edicao.html", tickets=tickets, ticket_sel=ticket_sel, record=record)


# ─── Produtividade ───────────────────────────────────────────────────────────

@gestao_bp.route("/produtividade")
@login_required
def produtividade():
    if not current_user.pode_produtividade and not current_user.is_admin:
        flash("Sem permissão.", "danger")
        return redirect(url_for("core.dashboard"))

    from collections import Counter
    retornos = _get_retornos_data()

    # Contagem por usuário (responsavel_execucao, avaliador_retorno, aprovador_retorno)
    exec_por_user = Counter()
    aval_por_user = Counter()
    valid_por_user = Counter()
    for r in retornos:
        exec_resp = r.get("responsavel_execucao", "").strip()
        aval = r.get("avaliador_retorno", "").strip()
        aprov = r.get("aprovador_retorno", "").strip()
        if exec_resp and exec_resp != "-":
            exec_por_user[exec_resp] += 1
        if aval and aval != "-":
            aval_por_user[aval] += 1
        if aprov and aprov != "-":
            valid_por_user[aprov] += 1

    # Totais por mês (últimos 6)
    from datetime import date
    from calendar import monthrange
    meses_pt = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    hoje = date.today()
    meses_dados = []
    for i in range(5, -1, -1):
        m = hoje.month - i
        a = hoje.year
        while m < 1: m += 12; a -= 1
        while m > 12: m -= 12; a += 1
        total_mes = sum(1 for r in retornos if parse_date(r.get("gp_recebimento_cliente",""))
                        and parse_date(r.get("gp_recebimento_cliente","")).month == m
                        and parse_date(r.get("gp_recebimento_cliente","")).year == a)
        meses_dados.append({"mes": f"{meses_pt[m-1]}/{a}", "total": total_mes})

    usuarios = sorted(set(list(exec_por_user.keys()) + list(aval_por_user.keys()) + list(valid_por_user.keys())))
    prod_data = []
    for u in usuarios:
        if exec_por_user[u] or aval_por_user[u] or valid_por_user[u]:
            prod_data.append({"nome": u, "exec": exec_por_user[u], "aval": aval_por_user[u], "valid": valid_por_user[u]})

    return render_template("produtividade.html", prod_data=prod_data, meses_dados=meses_dados)


# ─── Gráficos ────────────────────────────────────────────────────────────────

@gestao_bp.route("/graficos")
@login_required
def graficos():
    if not current_user.pode_graficos and not current_user.is_admin:
        flash("Sem permissão.", "danger")
        return redirect(url_for("core.dashboard"))

    from config import Config
    return render_template("graficos.html", setores=Config.SETORES_EXECUTORES)


@gestao_bp.route("/api/patrocinadores")
@login_required
def api_patrocinadores():
    from models import Patrocinador
    pats = Patrocinador.query.filter_by(ativo="Sim").order_by(Patrocinador.nome_patrocinador).all()
    return jsonify([p.nome_patrocinador for p in pats if p.nome_patrocinador])

@gestao_bp.route("/api/graficos")
@login_required
def api_graficos():
    retornos = _get_retornos_data()
    for r in retornos:
        r["_prazo"] = classificar_prazo(r)
    mes = request.args.get("mes", "Todos")
    ano = request.args.get("ano", "Todos")
    setor = request.args.get("setor", "Todos")

    def match(r):
        d = parse_date(r.get("gp_recebimento_cliente", ""))
        if d is None:
            return True
        if ano != "Todos" and d.year != int(ano):
            return False
        if mes != "Todos":
            meses_num = {"Jan": 1, "Fev": 2, "Mar": 3, "Abr": 4, "Mai": 5, "Jun": 6,
                         "Jul": 7, "Ago": 8, "Set": 9, "Out": 10, "Nov": 11, "Dez": 12}
            if d.month != meses_num.get(mes, 0):
                return False
        return True

    if setor != "Todos":
        retornos = [r for r in retornos if r.get("area_responsavel") == setor]
    retornos = [r for r in retornos if match(r)]

    from config import Config
    STATUS_FLUXO = Config.STATUS_FLUXO

    status_andamento = [s for s in STATUS_FLUXO if s != "Concluído"]
    donut_data = {}
    for s in status_andamento:
        n = sum(1 for r in retornos if r["status_retorno"] == s)
        if n > 0:
            donut_data[s] = n

    total = sum(1 for r in retornos if r.get("gp_recebimento_cliente", "").strip())
    avaliados = sum(1 for r in retornos if r.get("avaliador_retorno", "").strip())
    aprovados = sum(1 for r in retornos if r.get("status_ce", "").lower() == "aprovado")
    enviados = sum(1 for r in retornos if r.get("gp_envio_cliente", "").strip())

    from config import Config
    setor_counts = {}
    for s in Config.SETORES_EXECUTORES:
        n = sum(1 for r in retornos if r.get("area_responsavel") == s)
        if n > 0:
            setor_counts[s] = n

    # Prazo data (apenas enviados — calcula dias entre recebimento e envio)
    df_env = [r for r in retornos if r.get("gp_envio_cliente", "").strip()]
    dentro = 0
    atraso = 0
    for r in df_env:
        d_rec = parse_date(r.get("gp_recebimento_cliente", ""))
        d_env = parse_date(r.get("gp_envio_cliente", ""))
        d_custom = parse_date(r.get("prazo_execucao_customizado", ""))
        cls = prazo_envio(d_rec, d_env, d_custom)
        if cls in ("Dentro do Prazo", "Prazo Final", "Próximo ao Vencimento"):
            dentro += 1
        elif cls == "Em Atraso":
            atraso += 1

    # Lead time + cadastrados/enviados por mês (últimos 6 meses)
    meses_pt = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    hoje = date.today()
    # Gera meses em ordem cronológica crescente (ex: Dez/2025, Jan/2026...)
    meses_lista = []
    for i in range(5, -1, -1):
        m = hoje.month - i
        a = hoje.year
        while m < 1: m += 12; a -= 1
        while m > 12: m -= 12; a += 1
        meses_lista.append((m, a, f"{meses_pt[m-1]}/{a}"))

    # Inicializa arrays na ordem correta
    cad_arr = [0] * len(meses_lista)
    env_arr = [0] * len(meses_lista)
    lead_raw = [[] for _ in meses_lista]
    labels_arr = [chave for _, _, chave in meses_lista]

    # Mapa: label → índice
    label_idx = {chave: i for i, (_, _, chave) in enumerate(meses_lista)}

    for r in retornos:
        d_rec = parse_date(r.get("gp_recebimento_cliente", ""))
        d_env = parse_date(r.get("gp_envio_cliente", ""))
        if d_rec:
            chave = f"{meses_pt[d_rec.month-1]}/{d_rec.year}"
            idx = label_idx.get(chave)
            if idx is not None:
                cad_arr[idx] += 1
        if d_env:
            chave = f"{meses_pt[d_env.month-1]}/{d_env.year}"
            idx = label_idx.get(chave)
            if idx is not None:
                env_arr[idx] += 1
        if d_rec and d_env and d_env >= d_rec:
            chave = f"{meses_pt[d_rec.month-1]}/{d_rec.year}"
            idx = label_idx.get(chave)
            if idx is not None:
                lead_raw[idx].append((d_env - d_rec).days)

    lead_arr = [round(sum(v)/len(v), 1) if v else 0 for v in lead_raw]

    return jsonify(donut=donut_data, counters={"total": total, "avaliados": avaliados,
                   "aprovados": aprovados, "enviados": enviados},
                   setores=setor_counts, prazo={"Dentro do Prazo": dentro, "Em Atraso": atraso},
                   lead=lead_arr, lead_labels=labels_arr,
                   cadastrados_mensal=cad_arr, enviados_mensal=env_arr, line_labels=labels_arr)


# ─── Indicadores Diretoria ───────────────────────────────────────────────────

@gestao_bp.route("/indicadores-diretoria")
@login_required
def indicadores_diretoria():
    if not current_user.pode_diretoria:
        flash("Sem permissão.", "danger")
        return redirect(url_for("core.dashboard"))
    return render_template("indicadores_diretoria.html")
